"""Build Telco Churn manager report (Excel + Power BI-style HTML)."""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

OUT = Path(__file__).resolve().parent
data = json.loads((OUT / "dashboard_data.json").read_text(encoding="utf-8"))
if "metrics" not in data:
    data["metrics"] = {
        "accuracy": 73.17,
        "precision": 49.65,
        "recall": 75.67,
        "f1": 0.60,
        "auc": 0.81,
        "tn": 748,
        "fp": 287,
        "fn": 91,
        "tp": 283,
    }

k, m = data["kpis"], data["metrics"]

header_fill = PatternFill("solid", fgColor="0B3D5C")
accent_fill = PatternFill("solid", fgColor="0B6E4F")
white = Font(color="FFFFFF", bold=True, size=14)
title_font = Font(bold=True, size=16, color="0B3D5C")
kpi_font = Font(bold=True, size=18, color="0B3D5C")
thin = Border(
    left=Side(style="thin", color="D7E0EA"),
    right=Side(style="thin", color="D7E0EA"),
    top=Side(style="thin", color="D7E0EA"),
    bottom=Side(style="thin", color="D7E0EA"),
)


def style_header_row(ws, cols=8):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = white


wb = Workbook()

# ----- Page 1 -----
ws1 = wb.active
ws1.title = "01_Executive_Summary"
ws1["A1"] = "TELCO CHURN — MANAGER REPORT (Power BI pack)"
ws1.merge_cells("A1:H1")
style_header_row(ws1, 8)
ws1.row_dimensions[1].height = 28
ws1["A2"] = (
    "Page 1: Executive Summary. Present this sheet, or import Customers into Power BI Desktop and Save as .pbix."
)
ws1.merge_cells("A2:H2")

kpis = [
    ("Total Customers", f"{k['total']:,}", "E8EEF6"),
    ("Churn Rate", f"{k['churn_rate']}%", "FDECEC"),
    ("Churned Customers", f"{k['churned']:,}", "FFF4E8"),
    ("Est. Monthly Revenue at Risk", f"${k['revenue_at_risk']:,.0f}", "FFF4E8"),
    ("Accuracy", f"{m['accuracy']}%", "E8F5EF"),
    ("Recall (catch rate)", f"{m['recall']}%", "E8F5EF"),
    ("Precision", f"{m['precision']}%", "E8EEF6"),
    ("F1-score", f"{m['f1']}", "E8EEF6"),
]
for i, (label, val, color) in enumerate(kpis):
    col = i + 1
    fill = PatternFill("solid", fgColor=color)
    ws1.cell(4, col, label).font = Font(size=10, color="5B6B7C")
    ws1.cell(5, col, val).font = kpi_font
    ws1.cell(4, col).fill = fill
    ws1.cell(5, col).fill = fill
    ws1.cell(4, col).border = thin
    ws1.cell(5, col).border = thin

ws1["A7"] = "Key insight for managers"
ws1["A7"].font = title_font
ws1["A8"] = (
    "Contract type is the #1 churn driver. Month-to-month + short-tenure customers leave most. "
    "Model recall ~76% means most churners can be flagged early; use high-risk lists for save offers."
)
ws1.merge_cells("A8:H8")
ws1.row_dimensions[8].height = 40

ws1["A10"] = "Contract"
ws1["B10"] = "ChurnRatePct"
ws1["C10"] = "Customers"
for i, r in enumerate(data["by_contract"], start=11):
    ws1.cell(i, 1, r["Contract"])
    ws1.cell(i, 2, r["ChurnRatePct"])
    ws1.cell(i, 3, r["Customers"])

chart1 = BarChart()
chart1.type = "col"
chart1.title = "Churn Rate % by Contract"
chart1.y_axis.title = "Churn rate %"
n = len(data["by_contract"])
chart1.add_data(Reference(ws1, min_col=2, min_row=10, max_row=10 + n), titles_from_data=True)
chart1.set_categories(Reference(ws1, min_col=1, min_row=11, max_row=10 + n))
chart1.dataLabels = DataLabelList()
chart1.dataLabels.showVal = True
ws1.add_chart(chart1, "E10")

start = 11 + n + 2
ws1.cell(start, 1, "TenureBand")
ws1.cell(start, 2, "ChurnRatePct")
for i, r in enumerate(data["by_tenure"], start=start + 1):
    ws1.cell(i, 1, str(r["TenureBand"]))
    ws1.cell(i, 2, r["ChurnRatePct"])
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Churn Rate % by Tenure Band"
nt = len(data["by_tenure"])
chart2.add_data(Reference(ws1, min_col=2, min_row=start, max_row=start + nt), titles_from_data=True)
chart2.set_categories(Reference(ws1, min_col=1, min_row=start + 1, max_row=start + nt))
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showVal = True
ws1.add_chart(chart2, "E25")

for col in range(1, 9):
    ws1.column_dimensions[get_column_letter(col)].width = 18

# ----- Page 2 -----
ws2 = wb.create_sheet("02_Who_Is_Churning")
ws2["A1"] = "Page 2: Who Is Churning"
ws2.merge_cells("A1:F1")
style_header_row(ws2, 6)


def write_bar_table(ws, title, records, keys, start_row, chart_title, chart_anchor):
    ws.cell(start_row, 1, title).font = title_font
    for j, key in enumerate(keys, 1):
        ws.cell(start_row + 1, j, key).font = Font(bold=True)
    for i, r in enumerate(records):
        for j, key in enumerate(keys, 1):
            ws.cell(start_row + 2 + i, j, r.get(key, ""))
    cidx = keys.index("ChurnRatePct") + 1
    nrec = len(records)
    chart = BarChart()
    chart.type = "col"
    chart.title = chart_title
    chart.add_data(
        Reference(ws, min_col=cidx, min_row=start_row + 1, max_row=start_row + 1 + nrec),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + nrec))
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    ws.add_chart(chart, chart_anchor)
    return start_row + 3 + nrec


row = write_bar_table(
    ws2,
    "By Payment Method",
    data["by_payment"],
    ["PaymentMethod", "Customers", "Churned", "ChurnRatePct", "RevenueAtRisk"],
    3,
    "Churn % by Payment Method",
    "G3",
)
row = write_bar_table(
    ws2,
    "By Tech Support",
    data["by_tech"],
    ["TechSupport", "Customers", "Churned", "ChurnRatePct", "RevenueAtRisk"],
    row + 2,
    "Churn % by Tech Support",
    "G20",
)
write_bar_table(
    ws2,
    "By Risk Segment",
    data["by_risk"],
    ["RiskSegment", "Customers", "Churned", "ChurnRatePct", "RevenueAtRisk"],
    row + 2,
    "Churn % by Risk Segment",
    "G35",
)
for col in range(1, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 22

# ----- Page 3 -----
ws3 = wb.create_sheet("03_Model_Trust")
ws3["A1"] = "Page 3: Model Trust — Accuracy, Precision, Recall, F1, Confusion Matrix"
ws3.merge_cells("A1:E1")
style_header_row(ws3, 5)

ws3["A3"] = "Metric"
ws3["B3"] = "Value"
ws3["C3"] = "Manager meaning"
for c in range(1, 4):
    ws3.cell(3, c).font = Font(bold=True, color="FFFFFF")
    ws3.cell(3, c).fill = header_fill

metrics_rows = [
    ("Accuracy", f"{m['accuracy']}%", "Overall correct predictions (Churn or No Churn)"),
    ("Precision (Churn)", f"{m['precision']}%", "Of predicted churners, how many truly left"),
    ("Recall (Churn)", f"{m['recall']}%", "Of actual churners, how many we caught"),
    ("F1-score", f"{m['f1']}", "Balance of precision and recall"),
    ("ROC-AUC", f"{m['auc']}", "Ranking quality for prioritising outreach"),
]
for i, (a, b, c) in enumerate(metrics_rows, start=4):
    ws3.cell(i, 1, a)
    ws3.cell(i, 2, b)
    ws3.cell(i, 3, c)

ws3["A11"] = "Confusion Matrix"
ws3["A11"].font = title_font
ws3["B12"] = "Pred No Churn"
ws3["C12"] = "Pred Churn"
ws3["A13"] = "Actual No Churn"
ws3["B13"] = m["tn"]
ws3["C13"] = m["fp"]
ws3["A14"] = "Actual Churn"
ws3["B14"] = m["fn"]
ws3["C14"] = m["tp"]
for r in range(12, 15):
    for c in range(1, 4):
        ws3.cell(r, c).border = thin
        if r == 12 or c == 1:
            ws3.cell(r, c).font = Font(bold=True)

ws3["A16"] = "Interpretation"
ws3["A17"] = (
    f"TN={m['tn']} correctly retained; FP={m['fp']} false alarms; "
    f"FN={m['fn']} missed churners (costliest); TP={m['tp']} correctly caught. "
    "For retention, prioritise high recall even if precision is moderate."
)
ws3.merge_cells("A17:E17")
ws3.row_dimensions[17].height = 45
for col in range(1, 6):
    ws3.column_dimensions[get_column_letter(col)].width = 24

# ----- Page 4 -----
ws4 = wb.create_sheet("04_Action_Plan")
ws4["A1"] = "Page 4: Retention Action Plan"
ws4.merge_cells("A1:E1")
style_header_row(ws4, 5)
headers = ["Priority", "Segment", "ChurnInsight", "RecommendedAction", "Owner"]
for j, h in enumerate(headers, 1):
    cell = ws4.cell(3, j, h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = accent_fill
for i, a in enumerate(data["actions"], start=4):
    for j, h in enumerate(headers, 1):
        ws4.cell(i, j, a.get(h, ""))
        ws4.cell(i, j).alignment = Alignment(wrap_text=True, vertical="top")
        ws4.cell(i, j).border = thin
ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 32
ws4.column_dimensions["C"].width = 36
ws4.column_dimensions["D"].width = 55
ws4.column_dimensions["E"].width = 18
for r in range(4, 4 + len(data["actions"])):
    ws4.row_dimensions[r].height = 40

# Customers for Power BI
ws5 = wb.create_sheet("Customers")
cust = pd.read_csv(OUT / "Telco_Churn_PowerBI.csv")
for r_idx, row in enumerate(dataframe_to_rows(cust, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws5.cell(r_idx, c_idx, value)

xlsx_path = OUT / "Telco_Churn_Manager_Report.xlsx"
wb.save(xlsx_path)
print("Excel report:", xlsx_path)

# ----- HTML multi-page report -----
pages_actions = "".join(
    "<tr>"
    f"<td>{a['Priority']}</td><td>{a['Segment']}</td><td>{a['ChurnInsight']}</td>"
    f"<td>{a['RecommendedAction']}</td><td>{a['Owner']}</td>"
    "</tr>"
    for a in data["actions"]
)

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Telco Churn Manager Report — Power BI Style</title>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
<style>
:root {{ --navy:#0B3D5C; --accent:#0B6E4F; --warn:#C44900; --danger:#B42318; --bg:#F3F6F9; --card:#fff; --muted:#5B6B7C; --line:#D5DEE8; }}
* {{ box-sizing:border-box; }}
body {{ margin:0; font-family:"Segoe UI",Calibri,Arial,sans-serif; background:var(--bg); color:#1A2332; }}
.topbar {{ background:var(--navy); color:#fff; padding:14px 28px; display:flex; justify-content:space-between; align-items:center; gap:12px; flex-wrap:wrap; }}
.topbar h1 {{ margin:0; font-size:20px; }}
.topbar .sub {{ opacity:.85; font-size:12px; }}
.tabs {{ display:flex; gap:8px; padding:12px 28px; background:#E7EEF5; border-bottom:1px solid var(--line); flex-wrap:wrap; }}
.tab {{ border:1px solid var(--line); background:#fff; color:var(--navy); padding:10px 14px; border-radius:999px; cursor:pointer; font-weight:600; font-size:13px; }}
.tab.active {{ background:var(--navy); color:#fff; }}
.page {{ display:none; padding:22px 28px 40px; max-width:1220px; margin:0 auto; }}
.page.active {{ display:block; }}
.kpis {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-bottom:16px; }}
.kpi {{ background:var(--card); border:1px solid var(--line); border-radius:12px; padding:14px 16px; }}
.kpi .label {{ color:var(--muted); font-size:11px; text-transform:uppercase; letter-spacing:.04em; }}
.kpi .value {{ font-size:26px; font-weight:700; margin-top:6px; color:var(--navy); }}
.kpi.danger .value {{ color:var(--danger); }} .kpi.warn .value {{ color:var(--warn); }} .kpi.good .value {{ color:var(--accent); }}
.grid {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; }}
.card {{ background:var(--card); border:1px solid var(--line); border-radius:12px; padding:14px 16px; margin-bottom:14px; }}
.card h2 {{ margin:0 0 4px; font-size:16px; color:var(--navy); }}
.hint {{ margin:0 0 10px; color:var(--muted); font-size:13px; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th,td {{ padding:8px 6px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; }}
th {{ color:var(--muted); font-size:12px; }}
.insight {{ background:#FFF8F1; border:1px solid #F0D2B0; border-radius:12px; padding:14px 16px; margin-bottom:14px; }}
.cm {{ display:grid; grid-template-columns:repeat(2,1fr); gap:10px; max-width:420px; }}
.cm div {{ border:1px solid var(--line); border-radius:10px; padding:12px; text-align:center; background:#F8FAFC; }}
.cm b {{ display:block; font-size:22px; color:var(--navy); }}
@media (max-width:900px) {{ .kpis,.grid,.cm {{ grid-template-columns:1fr 1fr; }} }}
@media print {{ .tabs {{ display:none; }} .page {{ display:block !important; page-break-after:always; }} }}
</style>
</head>
<body>
<div class="topbar">
  <div>
    <h1>Telco Customer Churn — Manager Report</h1>
    <div class="sub">Power BI–style report · BDA601 · Accuracy / Precision / Recall / F1 / Confusion Matrix</div>
  </div>
  <div class="sub">Accuracy {m['accuracy']}% · Recall {m['recall']}% · Precision {m['precision']}%</div>
</div>
<div class="tabs">
  <button class="tab active" onclick="showPage(0)">1. Executive Summary</button>
  <button class="tab" onclick="showPage(1)">2. Who Is Churning</button>
  <button class="tab" onclick="showPage(2)">3. Model Trust</button>
  <button class="tab" onclick="showPage(3)">4. Action Plan</button>
</div>

<section class="page active" id="p0">
  <div class="kpis">
    <div class="kpi"><div class="label">Total Customers</div><div class="value">{k['total']:,}</div></div>
    <div class="kpi danger"><div class="label">Churn Rate</div><div class="value">{k['churn_rate']}%</div></div>
    <div class="kpi warn"><div class="label">Monthly Revenue at Risk</div><div class="value">${k['revenue_at_risk']:,.0f}</div></div>
    <div class="kpi good"><div class="label">High-Risk Customers</div><div class="value">{k['high_risk']:,}</div></div>
  </div>
  <div class="insight"><strong>Manager takeaway:</strong> Contract is the #1 driver. Focus retention on month-to-month + new customers (0–12 months). Model recall ~{m['recall']}% supports early outreach lists.</div>
  <div class="grid">
    <div class="card"><h2>Churn rate by contract</h2><p class="hint">Primary driver from the decision tree.</p><div id="c0a"></div></div>
    <div class="card"><h2>Churn by tenure band</h2><p class="hint">New customers are the priority window.</p><div id="c0b"></div></div>
  </div>
</section>

<section class="page" id="p1">
  <div class="grid">
    <div class="card"><h2>Payment method risk</h2><div id="c1a"></div></div>
    <div class="card"><h2>Tech support coverage</h2><div id="c1b"></div></div>
  </div>
  <div class="card"><h2>Risk segment mix</h2><p class="hint">Use High Risk for weekly win-back lists.</p><div id="c1c"></div></div>
</section>

<section class="page" id="p2">
  <div class="kpis">
    <div class="kpi"><div class="label">Accuracy</div><div class="value">{m['accuracy']}%</div></div>
    <div class="kpi"><div class="label">Precision</div><div class="value">{m['precision']}%</div></div>
    <div class="kpi good"><div class="label">Recall</div><div class="value">{m['recall']}%</div></div>
    <div class="kpi"><div class="label">F1-score</div><div class="value">{m['f1']}</div></div>
  </div>
  <div class="card">
    <h2>What these metrics mean</h2>
    <table>
      <tr><th>Metric</th><th>Meaning for managers</th></tr>
      <tr><td>Accuracy</td><td>Overall % of customers labelled correctly.</td></tr>
      <tr><td>Precision</td><td>Of predicted churners, how many truly left (false-alarm control).</td></tr>
      <tr><td>Recall</td><td>Of actual churners, how many we caught (coverage for save campaigns).</td></tr>
      <tr><td>F1-score</td><td>Balance between precision and recall.</td></tr>
      <tr><td>ROC-AUC ({m['auc']})</td><td>How well the model ranks high-risk customers for contact priority.</td></tr>
    </table>
  </div>
  <div class="card">
    <h2>Confusion matrix (test set)</h2>
    <div class="cm">
      <div><span>True Negatives</span><b>{m['tn']}</b><small>Correctly retained</small></div>
      <div><span>False Positives</span><b>{m['fp']}</b><small>False alarms</small></div>
      <div><span>False Negatives</span><b>{m['fn']}</b><small>Missed churners</small></div>
      <div><span>True Positives</span><b>{m['tp']}</b><small>Caught churners</small></div>
    </div>
  </div>
</section>

<section class="page" id="p3">
  <div class="card">
    <h2>Prioritised retention actions</h2>
    <p class="hint">Assign owners and track weekly.</p>
    <table>
      <thead><tr><th>#</th><th>Segment</th><th>Insight</th><th>Action</th><th>Owner</th></tr></thead>
      <tbody>{pages_actions}</tbody>
    </table>
  </div>
</section>

<script>
const DATA = {json.dumps(data)};
const layoutBase = {{margin:{{t:10,r:10,b:80,l:50}}, paper_bgcolor:'rgba(0,0,0,0)', plot_bgcolor:'rgba(0,0,0,0)', font:{{family:'Segoe UI', size:12}}, height:320}};
function bar(id, rows, x, y, color) {{
  Plotly.newPlot(id, [{{type:'bar', x:rows.map(r=>r[x]), y:rows.map(r=>r[y]), marker:{{color}}, text:rows.map(r=>r[y]+'%'), textposition:'auto'}}],
    Object.assign({{}}, layoutBase, {{yaxis:{{title:'Churn rate %'}}}}), {{displayModeBar:false, responsive:true}});
}}
function showPage(i) {{
  document.querySelectorAll('.page').forEach((p,idx)=>p.classList.toggle('active', idx===i));
  document.querySelectorAll('.tab').forEach((t,idx)=>t.classList.toggle('active', idx===i));
  setTimeout(()=>window.dispatchEvent(new Event('resize')), 50);
}}
bar('c0a', DATA.by_contract, 'Contract', 'ChurnRatePct', '#0B6E4F');
bar('c0b', DATA.by_tenure, 'TenureBand', 'ChurnRatePct', '#1F6FEB');
bar('c1a', DATA.by_payment, 'PaymentMethod', 'ChurnRatePct', '#C44900');
bar('c1b', DATA.by_tech, 'TechSupport', 'ChurnRatePct', '#7C3AED');
Plotly.newPlot('c1c', [{{type:'pie', labels:DATA.by_risk.map(r=>r.RiskSegment), values:DATA.by_risk.map(r=>r.Customers), hole:0.45,
  marker:{{colors:['#0B6E4F','#C44900','#B42318']}}, textinfo:'label+percent'}}], Object.assign({{}}, layoutBase, {{showlegend:false}}), {{displayModeBar:false, responsive:true}});
</script>
</body></html>
"""
(OUT / "Telco_Churn_Manager_Report.html").write_text(html, encoding="utf-8")
print("HTML report:", OUT / "Telco_Churn_Manager_Report.html")

(OUT / "README.md").write_text(
    """# Power BI Manager Report — Telco Churn

## Open the report now

1. **Interactive report:** `Telco_Churn_Manager_Report.html` (4 pages)
2. **Excel report (charts included):** `Telco_Churn_Manager_Report.xlsx`
   - `01_Executive_Summary`
   - `02_Who_Is_Churning`
   - `03_Model_Trust`
   - `04_Action_Plan`
   - `Customers` (import this into Power BI Desktop)

## Save as official `.pbix`

1. Open Power BI Desktop
2. Get data → Excel → `Telco_Churn_Manager_Report.xlsx` → load `Customers`
3. Build visuals using the 4 Excel pages as the layout guide
4. File → Save as → `Telco_Churn_Manager_Dashboard.pbix`

A binary `.pbix` can only be written by Power BI Desktop. This folder contains the finished manager report content.
""",
    encoding="utf-8",
)
print("Done")
